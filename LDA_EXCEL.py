# -*- coding: utf-8 -*-
# 读写2003 excel
import openpyxl
import xlwt
import xlrd
from xlutils.copy import copy
# 读写2007 excel



def write03Excel(path):
    book = xlwt.Workbook()
    sheet1 = book.add_sheet("2003测试表")
    value = [["1,1", "1,2", "1,3"],
             ["2,1", "2,2", "2,3"],
             ["3,1","3,2","3,3"]]
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet1.write(i, j, value[i][j])
    book.save(path)
    print("写入数据成功！")


def read03Excel(path):
    book = xlrd.open_workbook(path)
    sheet1 = book.sheet_names()
    worksheet = book.sheet_by_name(sheet1[0])
    for i in range(0, worksheet.nrows):
        row = worksheet.row(i)
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t", end="")
        print()

import openpyxl

def write07Excel(path):
    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = 'Sheet1'

    value = [["1,1", "1,2", "1,3"],
             ["2,1", "2,2", "2,3"],
             ["3,1","3,2","3,3"]]
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet1.cell(row=i+1, column=j+1, value=str(value[i][j]))

    book.save(path)
    print("写入数据成功！")


def read07Excel(path):
    book = openpyxl.load_workbook(path)
    sheet1 = book.get_sheet_by_name('Sheet1')

    for row in sheet1.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

def excelwrite(filename,value):         
    workbook = xlrd.open_workbook(filename) 
    sheet = workbook.sheet_by_index(0)  
    rowNum = sheet.nrows    
    colNum = sheet.ncols    
    newbook = copy(workbook)    
    newsheet = newbook.get_sheet(0) # 在末尾增加新行
    for i in range(0,len(value)):
        newsheet.write(rowNum, i, value[i]) # 覆盖保存  
    newbook.save(filename)

def add07Excel(path,value):
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']
    #for x in data:
    ws.append(value)
    wb.save(path)
    print("写入成功")

file_2003 = 'test/2003.xls'
file_2007 = 'test/2007.xlsx'

write03Excel(file_2003)
read03Excel(file_2003)

write07Excel(file_2007)
read07Excel(file_2007)